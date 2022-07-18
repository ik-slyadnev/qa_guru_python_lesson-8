import os
import shutil
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

# Создаем папку и архив
os.mkdir('resources')
zip_file = zipfile.ZipFile('resources\Архив содержимого.zip', 'w')


# Добавляем файлы в архив
def test_zip_files():
    for filename in os.listdir('folder'):
        zip_file.write(os.path.join('folder', filename), compress_type=zipfile.ZIP_DEFLATED)
    zip_file.close()


# Проверка файлов
def test_check_files():
    for file_info in zip_file.infolist():
        print(file_info.filename, file_info.date_time, file_info.file_size)

        zip_file.close()

# Разархивируем файл
def test_unzip_files():
    zip_file = zipfile.ZipFile('resources\Архив содержимого.zip')
    zip_file.extractall('resources/')
    zip_file.close()

# Проверка pdf
def test_read_pdf():
    reader = PdfReader('folder/Штирнер, Ницше Этика эгоизма. Нет ничего выше меня.pdf')
    page = reader.pages[1]
    text = page.extract_text()
    assert 'Этика эгоизма' in text

# Проверка xlsx
def test_read_xlsx():
    workbook = load_workbook('folder/file_excel.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=2, column=1).value
    assert 'Ivan' == name

# Проверка csv
def test_read_csv_():
    with open('folder/users.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'First_name' in str(headers)

# Удаляем папку
def test_remove_folder():
    shutil.rmtree('resources')